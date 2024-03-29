from flask import Blueprint, request, render_template, Response, send_file,session,current_app
import os
import csv
import io
import subprocess
from datetime import datetime
from openpyxl import Workbook,load_workbook
from openpyxl.styles import PatternFill,Font, Alignment,Border,Side

# Create a Blueprint object for Tab 4
tab4_bp = Blueprint('tab4', __name__, url_prefix='/tab4')
def create_iNMS_template(data):
    # Create a new workbook
    wb = Workbook()
    sheet = wb.active
    headers = [
        "S/N",
        "Group Name (My Network)",
        "Hostname",
        "IP Address",
        "SNMP Version (SNMPv1 is not allowed)",
        "Credentials for SNMPv2"
    ]
    sheet.append(headers)
    header_row=sheet[1]

    fill = PatternFill(start_color="99CC00", end_color="99CC00", fill_type="solid")
    font = Font(bold=True, name='Arial',size=10)
    font_text=Font(name='Arial',size=10)
    alignment = Alignment(horizontal='center')
    border = Border(bottom=Side(border_style="thin"))
    for cell in header_row:
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment
        cell.border= border
    # Write data to the worksheet
    series_number = 1
    for item in data:
        # Check if the length of the item is at least 2
        if len(item) == 2:
            # Extract the necessary information from the item
            group_name = "PCTN - ATN950D" if "ATN950D" in item[0] else "PCTN - ATN910CG"
            hostname = item[0]
            ip_address = item[1]

            # Append a new row with the extracted information
            row_data = [series_number, group_name, hostname, ip_address, 'SNMPv2', 'PCTN-read']
            sheet.append(row_data)

            series_number += 1
        else:
            # Create a separate row for errors
            error_row = ['Error'] * len(headers)
            sheet.append(error_row)
                # Apply borders to the row
        for cell in sheet[sheet.max_row]:
            cell.border = border
            cell.font=font_text
    return wb

def read_data_from_file(file_path):
    with open(file_path, 'r') as file:
        data = [line.strip().split(',') for line in file]
    return data
@tab4_bp.route('/', methods=['GET', 'POST'])
def tab4_home():
    if request.method == 'POST':
        input_text = request.form['NE']

        # Modify the input.txt file
        with open('input.txt', 'w') as file:
            file.write(input_text)

        #script_path = '/home/rancid/test/app_ise.sh'
        #subprocess.run(['bash', script_path])

        # Read data from a file (in this case, a CSV file)
        output_file_path = 'C:/Users/sutha/OneDrive/Desktop/blueprint/downloads/output.txt'
        data = read_data_from_file(output_file_path)

        # Create the Excel template with data from the file
        wb = create_iNMS_template(data)

        # Specify the file name with a dynamic file name
        current_datetime_str = datetime.now().strftime("%Y%m%d")
        output_file_name = 'NMS_V2.0_{}.xlsx'.format(current_datetime_str)
        output_folder='C:/Users/sutha/OneDrive/Desktop/blueprint/downloads/'
        
        # Combine the folder path and file name to get the full output file path
        output_file_path = os.path.join(output_folder, output_file_name)

        # Save the workbook to the specified folder
        wb.save(output_file_path)

        download_successful = 'iNMS Template downloaded Successfully. Please check if all the fields are correct before raising the Surf+ Request'

        # Return the file as a response to the client
        return send_file(
            output_file_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=output_file_name
        )
      
    return render_template('tab4.html')